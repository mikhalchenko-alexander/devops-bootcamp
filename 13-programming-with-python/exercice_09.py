import openpyxl

employees_file = openpyxl.load_workbook("employees.xlsx")
employees_sheet = employees_file["Sheet1"]

employees_list = []
for row in range(2, employees_sheet.max_row + 1):
    employees_list.append({
        "name": employees_sheet.cell(row, 1).value,
        "years_of_experience": employees_sheet.cell(row, 2).value,
    })

employees_sorted = sorted(employees_list, key=lambda employee: employee["years_of_experience"])

employees_sorted_workbook = openpyxl.workbook.Workbook()
employees_sorted_sheet = employees_sorted_workbook.worksheets[0]
employees_sorted_sheet.title = "Employees Sorted"
employees_sorted_sheet.append(["Name", "Years of Experience"])

for employee in employees_sorted:
    employees_sorted_sheet.append([employee["name"], employee["years_of_experience"]])

employees_sorted_workbook.save("employees_sorted.xlsx")
